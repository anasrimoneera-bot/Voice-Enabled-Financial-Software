# -*- coding: utf-8 -*-
"""
财务语音记账 - 多人协同云服务
端口：5600
风格对齐 app.py（蓝鲸科技授权服务）：Flask + waitress + SQLite + session 鉴权

核心特性：
  - 用户登录（支持多人协作，admin 可在管理页加账号）
  - 手机网页 APP，支持语音录入、AI 解析、实时同步
  - 数据持久化到 SQLite，每条记录带"录入人"
  - 一键导出 Excel，列结构对齐用户提供的财务日记账模板
  - AI 识别在服务器端进行（前端只传原始语音文字）
"""
import os
import json
import sqlite3
import datetime
import hashlib
import re
import secrets
import tempfile
import threading
import time
import queue
from io import BytesIO
from flask import (
    Flask, request, jsonify, render_template,
    session, redirect, url_for, send_file, Response, make_response
)

# ============ 基础配置 ============
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, 'finance.db')
USERS_FILE = os.path.join(BASE_DIR, 'users.json')
AI_CONFIG_FILE = os.path.join(BASE_DIR, 'ai_config.json')

app = Flask(__name__)
app.secret_key = b"Finance_Voice_Journal_2026_SecretKey_Change_Me!"
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# HTTPS 部署后改 True；Nginx 反代 HTTP 也可以保持 False，因为 cookie 不出 LAN
app.config['SESSION_COOKIE_SECURE'] = False
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=30)


# ============ 工具函数 ============
def now_str():
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def get_client_ip():
    """优先 X-Forwarded-For（穿透 Nginx），其次 remote_addr。"""
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr or ''


def _atomic_write_json(path, data):
    """先写 tmp 再 os.replace，防止宕机时半截 JSON。"""
    d = os.path.dirname(path) or '.'
    fd, tmp = tempfile.mkstemp(prefix='.tmp_', suffix='.json', dir=d)
    try:
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.flush()
            try:
                os.fsync(f.fileno())
            except Exception:
                pass
        os.replace(tmp, path)
    except Exception:
        try:
            os.unlink(tmp)
        except Exception:
            pass
        raise


def hash_password(pwd, salt=None):
    """密码哈希：sha256(salt + pwd)，salt 16 字节 hex。"""
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.sha256((salt + pwd).encode('utf-8')).hexdigest()
    return f"{salt}${h}"


def verify_password(pwd, stored):
    """对比 hash_password 输出的格式 salt$hash。"""
    try:
        salt, _ = stored.split('$', 1)
    except ValueError:
        return False
    return hash_password(pwd, salt) == stored


# ============ 用户系统（JSON 存储） ============
_users_lock = threading.Lock()


def _load_users():
    if not os.path.exists(USERS_FILE):
        # 首次启动：创建默认 admin 账号
        default = {
            'admin': {
                'password': hash_password('admin123'),
                'display_name': '管理员',
                'role': 'admin',
                'created_at': now_str(),
                'enabled': True,
            }
        }
        _atomic_write_json(USERS_FILE, default)
        return default
    try:
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_users(users):
    _atomic_write_json(USERS_FILE, users)


def get_user(username):
    users = _load_users()
    return users.get(username)


# ============ AI 配置 ============
_ai_config_lock = threading.Lock()


def _load_ai_config():
    if not os.path.exists(AI_CONFIG_FILE):
        default = {'provider': 'deepseek', 'apikey': ''}
        _atomic_write_json(AI_CONFIG_FILE, default)
        return default
    try:
        with open(AI_CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'provider': 'deepseek', 'apikey': ''}


def _save_ai_config(cfg):
    _atomic_write_json(AI_CONFIG_FILE, cfg)


# ============ 数据库 ============
def get_db():
    conn = sqlite3.connect(DB_FILE, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL;')  # 多读单写并发更好
    return conn


def init_db():
    conn = get_db()
    try:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS entries (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            date          TEXT    NOT NULL,         -- YYYY/M/D
            year_month    TEXT    NOT NULL,         -- YYYY-MM 用于筛选
            summary       TEXT    NOT NULL,
            type          TEXT    NOT NULL CHECK(type IN ('income','expense')),
            amount        REAL    NOT NULL,
            remark        TEXT    DEFAULT '',
            raw_text      TEXT    DEFAULT '',       -- 原始语音文字
            creator       TEXT    NOT NULL,         -- 录入人 username
            creator_name  TEXT    NOT NULL,         -- 录入人显示名
            created_at    TEXT    NOT NULL,
            updated_at    TEXT    NOT NULL,
            deleted       INTEGER DEFAULT 0
        );
        CREATE INDEX IF NOT EXISTS idx_entries_ym ON entries(year_month);
        CREATE INDEX IF NOT EXISTS idx_entries_date ON entries(date);
        CREATE INDEX IF NOT EXISTS idx_entries_creator ON entries(creator);

        CREATE TABLE IF NOT EXISTS audit_log (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_id   INTEGER,
            action     TEXT NOT NULL,                -- create/update/delete/lock/unlock/import
            actor      TEXT NOT NULL,
            actor_name TEXT NOT NULL,
            detail     TEXT,
            created_at TEXT NOT NULL
        );

        -- 账套表：不同公司
        CREATE TABLE IF NOT EXISTS account_books (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            code        TEXT NOT NULL UNIQUE,        -- 账套代号（英文/数字，url-safe）
            name        TEXT NOT NULL,               -- 中文显示名（如"上海希联电子"）
            description TEXT DEFAULT '',
            enabled     INTEGER DEFAULT 1,
            sort_order  INTEGER DEFAULT 0,
            created_at  TEXT NOT NULL
        );

        -- 月度锁定表
        CREATE TABLE IF NOT EXISTS month_locks (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            book_code     TEXT NOT NULL,
            year_month    TEXT NOT NULL,             -- YYYY-MM
            locked_by     TEXT NOT NULL,
            locked_name   TEXT NOT NULL,
            locked_at     TEXT NOT NULL,
            note          TEXT DEFAULT '',
            UNIQUE(book_code, year_month)
        );
        CREATE INDEX IF NOT EXISTS idx_locks_ym ON month_locks(book_code, year_month);

        -- 类别表：按账套维度
        CREATE TABLE IF NOT EXISTS categories (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            book_code   TEXT NOT NULL,               -- 哪个账套的类别
            name        TEXT NOT NULL,
            type_hint   TEXT DEFAULT '',             -- 'income' / 'expense' / '' 偏好
            sort_order  INTEGER DEFAULT 0,
            enabled     INTEGER DEFAULT 1,
            UNIQUE(book_code, name)
        );

        -- 账号-账套授权（admin 默认能看全部）
        CREATE TABLE IF NOT EXISTS user_book_access (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            username  TEXT NOT NULL,
            book_code TEXT NOT NULL,
            UNIQUE(username, book_code)
        );
        CREATE INDEX IF NOT EXISTS idx_uba_user ON user_book_access(username);
        """)

        # 迁移：给 entries 加 book_code / category 列（如果还没有）
        cols = [r[1] for r in conn.execute("PRAGMA table_info(entries)").fetchall()]
        if 'book_code' not in cols:
            conn.execute("ALTER TABLE entries ADD COLUMN book_code TEXT NOT NULL DEFAULT 'default'")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_entries_book ON entries(book_code)")
        if 'category' not in cols:
            conn.execute("ALTER TABLE entries ADD COLUMN category TEXT DEFAULT ''")

        # 确保默认账套存在
        row = conn.execute("SELECT id FROM account_books WHERE code='default'").fetchone()
        if not row:
            conn.execute(
                "INSERT INTO account_books (code, name, description, sort_order, created_at) "
                "VALUES ('default', '默认账套', '系统自动创建', 0, ?)",
                (now_str(),)
            )

        # 给默认账套播种 8 个常用类别
        seed_categories = [
            ('餐饮', 'expense'), ('交通', 'expense'), ('办公', 'expense'),
            ('货款', 'expense'), ('工资', 'expense'), ('房租', 'expense'),
            ('水电', 'expense'), ('其他', ''),
        ]
        for idx, (name, hint) in enumerate(seed_categories):
            conn.execute(
                "INSERT OR IGNORE INTO categories (book_code, name, type_hint, sort_order) "
                "VALUES ('default', ?, ?, ?)",
                (name, hint, idx)
            )

        conn.commit()
    finally:
        conn.close()


def log_action(entry_id, action, actor, actor_name, detail=''):
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO audit_log (entry_id, action, actor, actor_name, detail, created_at) "
            "VALUES (?,?,?,?,?,?)",
            (entry_id, action, actor, actor_name, detail, now_str())
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print('[audit_log] 写入失败:', e)


# ============ 账套 & 锁定 & 类别 辅助函数 ============
def list_books(only_enabled=True):
    conn = get_db()
    try:
        sql = "SELECT * FROM account_books"
        if only_enabled:
            sql += " WHERE enabled=1"
        sql += " ORDER BY sort_order ASC, id ASC"
        return [dict(r) for r in conn.execute(sql).fetchall()]
    finally:
        conn.close()


def get_book(code):
    conn = get_db()
    try:
        r = conn.execute("SELECT * FROM account_books WHERE code=?", (code,)).fetchone()
        return dict(r) if r else None
    finally:
        conn.close()


def user_accessible_books(username):
    """返回该用户能访问的账套 code 列表。admin 看全部。"""
    user = get_user(username)
    if not user:
        return []
    if user.get('role') == 'admin':
        return [b['code'] for b in list_books()]
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT uba.book_code FROM user_book_access uba "
            "JOIN account_books ab ON ab.code = uba.book_code "
            "WHERE uba.username=? AND ab.enabled=1",
            (username,)
        ).fetchall()
        return [r['book_code'] for r in rows]
    finally:
        conn.close()


def can_access_book(username, book_code):
    return book_code in user_accessible_books(username)


def is_month_locked(book_code, year_month):
    conn = get_db()
    try:
        r = conn.execute(
            "SELECT id FROM month_locks WHERE book_code=? AND year_month=?",
            (book_code, year_month)
        ).fetchone()
        return r is not None
    finally:
        conn.close()


def get_month_lock(book_code, year_month):
    """返回锁定详情或 None"""
    conn = get_db()
    try:
        r = conn.execute(
            "SELECT * FROM month_locks WHERE book_code=? AND year_month=?",
            (book_code, year_month)
        ).fetchone()
        return dict(r) if r else None
    finally:
        conn.close()


# ============ 鉴权装饰器 ============
from functools import wraps


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if 'username' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'ok': False, 'error': 'unauthorized'}), 401
            return redirect(url_for('login_page'))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if 'username' not in session:
            return jsonify({'ok': False, 'error': 'unauthorized'}), 401
        user = get_user(session['username'])
        if not user or user.get('role') != 'admin':
            return jsonify({'ok': False, 'error': 'forbidden'}), 403
        return fn(*args, **kwargs)
    return wrapper


# ============ 页面路由 ============
@app.route('/')
def root():
    if 'username' in session:
        return redirect(url_for('app_page'))
    return redirect(url_for('login_page'))


@app.route('/login')
def login_page():
    if 'username' in session:
        return redirect(url_for('app_page'))
    return render_template('login.html')


@app.route('/app')
@login_required
def app_page():
    user = get_user(session['username'])
    return render_template(
        'app.html',
        username=session['username'],
        display_name=user.get('display_name', session['username']),
        is_admin=(user.get('role') == 'admin')
    )


@app.route('/admin')
@login_required
def admin_page():
    user = get_user(session['username'])
    if user.get('role') != 'admin':
        return "需要管理员权限", 403
    return render_template('admin.html', username=session['username'])


@app.route('/reports')
@login_required
def reports_page():
    user = get_user(session['username'])
    return render_template(
        'reports.html',
        username=session['username'],
        display_name=user.get('display_name', session['username']),
        is_admin=(user.get('role') == 'admin')
    )


# ============ 登录/登出 API ============
@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    if not username or not password:
        return jsonify({'ok': False, 'error': '用户名或密码为空'}), 400
    user = get_user(username)
    if not user or not user.get('enabled', True):
        return jsonify({'ok': False, 'error': '账号不存在或已禁用'}), 401
    if not verify_password(password, user['password']):
        return jsonify({'ok': False, 'error': '密码错误'}), 401
    session.permanent = True
    session['username'] = username
    return jsonify({
        'ok': True,
        'username': username,
        'display_name': user.get('display_name', username),
        'role': user.get('role', 'user')
    })


@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.pop('username', None)
    return jsonify({'ok': True})


@app.route('/api/me', methods=['GET'])
@login_required
def api_me():
    user = get_user(session['username'])
    books_codes = user_accessible_books(session['username'])
    all_books = list_books()
    accessible = [b for b in all_books if b['code'] in books_codes]
    return jsonify({
        'ok': True,
        'username': session['username'],
        'display_name': user.get('display_name', session['username']),
        'role': user.get('role', 'user'),
        'books': accessible,
    })


@app.route('/api/change_password', methods=['POST'])
@login_required
def api_change_password():
    data = request.get_json(force=True, silent=True) or {}
    old_pwd = data.get('old_password') or ''
    new_pwd = data.get('new_password') or ''
    if not new_pwd or len(new_pwd) < 6:
        return jsonify({'ok': False, 'error': '新密码至少 6 位'}), 400
    with _users_lock:
        users = _load_users()
        u = users.get(session['username'])
        if not u or not verify_password(old_pwd, u['password']):
            return jsonify({'ok': False, 'error': '原密码错误'}), 401
        u['password'] = hash_password(new_pwd)
        _save_users(users)
    return jsonify({'ok': True})


# ============ 用户管理 API（仅 admin） ============
@app.route('/api/users', methods=['GET'])
@admin_required
def api_list_users():
    users = _load_users()
    out = []
    for name, u in users.items():
        out.append({
            'username': name,
            'display_name': u.get('display_name', name),
            'role': u.get('role', 'user'),
            'enabled': u.get('enabled', True),
            'created_at': u.get('created_at', ''),
        })
    return jsonify({'ok': True, 'users': out})


@app.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get('username') or '').strip()
    display_name = (data.get('display_name') or '').strip() or username
    password = data.get('password') or ''
    role = data.get('role') or 'user'
    if not username or not re.match(r'^[A-Za-z0-9_\-]{2,32}$', username):
        return jsonify({'ok': False, 'error': '用户名只能是字母/数字/下划线/短横,2-32位'}), 400
    if not password or len(password) < 6:
        return jsonify({'ok': False, 'error': '密码至少 6 位'}), 400
    if role not in ('admin', 'user'):
        return jsonify({'ok': False, 'error': '角色非法'}), 400
    with _users_lock:
        users = _load_users()
        if username in users:
            return jsonify({'ok': False, 'error': '用户名已存在'}), 409
        users[username] = {
            'password': hash_password(password),
            'display_name': display_name,
            'role': role,
            'enabled': True,
            'created_at': now_str(),
        }
        _save_users(users)
    return jsonify({'ok': True})


@app.route('/api/users/<username>', methods=['PATCH'])
@admin_required
def api_update_user(username):
    data = request.get_json(force=True, silent=True) or {}
    with _users_lock:
        users = _load_users()
        if username not in users:
            return jsonify({'ok': False, 'error': '用户不存在'}), 404
        u = users[username]
        if 'display_name' in data:
            u['display_name'] = (data['display_name'] or '').strip() or username
        if 'enabled' in data:
            u['enabled'] = bool(data['enabled'])
        if 'role' in data and data['role'] in ('admin', 'user'):
            u['role'] = data['role']
        if 'password' in data and data['password']:
            if len(data['password']) < 6:
                return jsonify({'ok': False, 'error': '密码至少 6 位'}), 400
            u['password'] = hash_password(data['password'])
        _save_users(users)
    return jsonify({'ok': True})


@app.route('/api/users/<username>', methods=['DELETE'])
@admin_required
def api_delete_user(username):
    if username == session['username']:
        return jsonify({'ok': False, 'error': '不能删除自己'}), 400
    with _users_lock:
        users = _load_users()
        if username not in users:
            return jsonify({'ok': False, 'error': '用户不存在'}), 404
        del users[username]
        _save_users(users)
    # 同时清理该用户的账套授权
    conn = get_db()
    try:
        conn.execute("DELETE FROM user_book_access WHERE username=?", (username,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/users/<username>/books', methods=['GET'])
@admin_required
def api_get_user_books(username):
    """获取某用户授权的账套列表（admin 视角）"""
    if not get_user(username):
        return jsonify({'ok': False, 'error': '用户不存在'}), 404
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT book_code FROM user_book_access WHERE username=?",
            (username,)
        ).fetchall()
        return jsonify({'ok': True, 'book_codes': [r['book_code'] for r in rows]})
    finally:
        conn.close()


@app.route('/api/users/<username>/books', methods=['POST'])
@admin_required
def api_set_user_books(username):
    """设置某用户授权的账套（覆盖式）"""
    if not get_user(username):
        return jsonify({'ok': False, 'error': '用户不存在'}), 404
    data = request.get_json(force=True, silent=True) or {}
    book_codes = data.get('book_codes') or []
    if not isinstance(book_codes, list):
        return jsonify({'ok': False, 'error': 'book_codes 必须是数组'}), 400
    all_codes = set(b['code'] for b in list_books(only_enabled=False))
    invalid = [c for c in book_codes if c not in all_codes]
    if invalid:
        return jsonify({'ok': False, 'error': f'不存在的账套: {invalid}'}), 400
    conn = get_db()
    try:
        conn.execute("DELETE FROM user_book_access WHERE username=?", (username,))
        for code in book_codes:
            conn.execute(
                "INSERT INTO user_book_access (username, book_code) VALUES (?,?)",
                (username, code)
            )
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


# ============ 账套管理 API ============
@app.route('/api/books', methods=['GET'])
@login_required
def api_list_books():
    """普通用户：只返回有权访问的；admin：全部"""
    user = get_user(session['username'])
    all_books = list_books()
    if user.get('role') == 'admin':
        return jsonify({'ok': True, 'books': all_books, 'is_admin': True})
    codes = set(user_accessible_books(session['username']))
    return jsonify({
        'ok': True,
        'books': [b for b in all_books if b['code'] in codes],
        'is_admin': False,
    })


@app.route('/api/books', methods=['POST'])
@admin_required
def api_create_book():
    data = request.get_json(force=True, silent=True) or {}
    code = (data.get('code') or '').strip()
    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    if not code or not re.match(r'^[A-Za-z0-9_\-]{2,32}$', code):
        return jsonify({'ok': False, 'error': '账套代号必须是 2-32 位字母/数字/下划线'}), 400
    if not name:
        return jsonify({'ok': False, 'error': '账套名称不能为空'}), 400
    conn = get_db()
    try:
        row = conn.execute("SELECT id FROM account_books WHERE code=?", (code,)).fetchone()
        if row:
            return jsonify({'ok': False, 'error': '账套代号已存在'}), 409
        # sort_order 取当前最大值 + 1
        mx = conn.execute("SELECT MAX(sort_order) as m FROM account_books").fetchone()
        sort_order = (mx['m'] or 0) + 1
        conn.execute(
            "INSERT INTO account_books (code, name, description, sort_order, created_at) "
            "VALUES (?,?,?,?,?)",
            (code, name, description, sort_order, now_str())
        )
        # 给新账套自动播种 8 个默认类别
        seed = [('餐饮','expense'),('交通','expense'),('办公','expense'),
                ('货款','expense'),('工资','expense'),('房租','expense'),
                ('水电','expense'),('其他','')]
        for idx, (n, h) in enumerate(seed):
            conn.execute(
                "INSERT OR IGNORE INTO categories (book_code, name, type_hint, sort_order) "
                "VALUES (?,?,?,?)", (code, n, h, idx))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/books/<code>', methods=['PATCH'])
@admin_required
def api_update_book(code):
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM account_books WHERE code=?", (code,)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': '账套不存在'}), 404
        new_name = (data.get('name') or row['name']).strip()
        new_desc = data.get('description', row['description'])
        new_enabled = int(bool(data['enabled'])) if 'enabled' in data else row['enabled']
        if not new_name:
            return jsonify({'ok': False, 'error': '名称不能为空'}), 400
        conn.execute(
            "UPDATE account_books SET name=?, description=?, enabled=? WHERE code=?",
            (new_name, new_desc, new_enabled, code)
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/books/<code>', methods=['DELETE'])
@admin_required
def api_delete_book(code):
    """删除账套（仅当无数据时允许，否则建议禁用）"""
    if code == 'default':
        return jsonify({'ok': False, 'error': '默认账套不能删除'}), 400
    conn = get_db()
    try:
        cnt = conn.execute(
            "SELECT COUNT(*) as c FROM entries WHERE book_code=? AND deleted=0",
            (code,)
        ).fetchone()['c']
        if cnt > 0:
            return jsonify({
                'ok': False,
                'error': f'该账套有 {cnt} 条流水，不能删除（可禁用）'
            }), 400
        conn.execute("DELETE FROM categories WHERE book_code=?", (code,))
        conn.execute("DELETE FROM user_book_access WHERE book_code=?", (code,))
        conn.execute("DELETE FROM month_locks WHERE book_code=?", (code,))
        conn.execute("DELETE FROM account_books WHERE code=?", (code,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


# ============ 类别管理 API ============
@app.route('/api/categories', methods=['GET'])
@login_required
def api_list_categories():
    book_code = request.args.get('book_code') or 'default'
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM categories WHERE book_code=? AND enabled=1 "
            "ORDER BY sort_order ASC, id ASC",
            (book_code,)
        ).fetchall()
        return jsonify({'ok': True, 'categories': [dict(r) for r in rows]})
    finally:
        conn.close()


@app.route('/api/categories', methods=['POST'])
@admin_required
def api_create_category():
    data = request.get_json(force=True, silent=True) or {}
    book_code = (data.get('book_code') or 'default').strip()
    name = (data.get('name') or '').strip()
    type_hint = data.get('type_hint') or ''
    if not name:
        return jsonify({'ok': False, 'error': '类别名不能为空'}), 400
    if type_hint and type_hint not in ('income', 'expense'):
        return jsonify({'ok': False, 'error': 'type_hint 非法'}), 400
    if not get_book(book_code):
        return jsonify({'ok': False, 'error': '账套不存在'}), 404
    conn = get_db()
    try:
        try:
            mx = conn.execute(
                "SELECT MAX(sort_order) as m FROM categories WHERE book_code=?",
                (book_code,)
            ).fetchone()
            sort_order = (mx['m'] or 0) + 1
            conn.execute(
                "INSERT INTO categories (book_code, name, type_hint, sort_order) "
                "VALUES (?,?,?,?)",
                (book_code, name, type_hint, sort_order)
            )
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({'ok': False, 'error': '该账套下已存在同名类别'}), 409
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/categories/<int:cat_id>', methods=['PATCH'])
@admin_required
def api_update_category(cat_id):
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM categories WHERE id=?", (cat_id,)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': '类别不存在'}), 404
        new_name = (data.get('name') or row['name']).strip()
        new_hint = data.get('type_hint', row['type_hint']) or ''
        new_enabled = int(bool(data['enabled'])) if 'enabled' in data else row['enabled']
        if new_hint and new_hint not in ('income', 'expense'):
            return jsonify({'ok': False, 'error': 'type_hint 非法'}), 400
        conn.execute(
            "UPDATE categories SET name=?, type_hint=?, enabled=? WHERE id=?",
            (new_name, new_hint, new_enabled, cat_id)
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
@admin_required
def api_delete_category(cat_id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM categories WHERE id=?", (cat_id,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})


# ============ 月度锁定 API ============
@app.route('/api/locks', methods=['GET'])
@login_required
def api_list_locks():
    """列出某账套下所有已锁定的月份"""
    book_code = request.args.get('book_code') or 'default'
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM month_locks WHERE book_code=? ORDER BY year_month DESC",
            (book_code,)
        ).fetchall()
        return jsonify({'ok': True, 'locks': [dict(r) for r in rows]})
    finally:
        conn.close()


@app.route('/api/locks', methods=['POST'])
@login_required
def api_create_lock():
    """锁定某月（仅 admin）"""
    user = get_user(session['username'])
    if user.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '仅管理员可锁定月份'}), 403
    data = request.get_json(force=True, silent=True) or {}
    book_code = (data.get('book_code') or 'default').strip()
    year_month = (data.get('year_month') or '').strip()
    note = (data.get('note') or '').strip()
    if not re.match(r'^\d{4}-\d{2}$', year_month):
        return jsonify({'ok': False, 'error': 'year_month 格式应为 YYYY-MM'}), 400
    if not get_book(book_code):
        return jsonify({'ok': False, 'error': '账套不存在'}), 404
    conn = get_db()
    try:
        try:
            conn.execute(
                "INSERT INTO month_locks (book_code, year_month, locked_by, "
                "locked_name, locked_at, note) VALUES (?,?,?,?,?,?)",
                (book_code, year_month, session['username'],
                 user.get('display_name', session['username']),
                 now_str(), note)
            )
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({'ok': False, 'error': '该月已锁定'}), 409
    finally:
        conn.close()
    log_action(None, 'lock', session['username'],
               user.get('display_name', session['username']),
               f"账套={book_code} 月份={year_month} 备注={note}")
    _broadcast({'event': 'lock', 'book_code': book_code, 'year_month': year_month})
    return jsonify({'ok': True})


@app.route('/api/locks', methods=['DELETE'])
@login_required
def api_delete_lock():
    """解锁某月（仅 admin）"""
    user = get_user(session['username'])
    if user.get('role') != 'admin':
        return jsonify({'ok': False, 'error': '仅管理员可解锁'}), 403
    book_code = (request.args.get('book_code') or 'default').strip()
    year_month = (request.args.get('year_month') or '').strip()
    note = (request.args.get('note') or '').strip()
    conn = get_db()
    try:
        cur = conn.execute(
            "DELETE FROM month_locks WHERE book_code=? AND year_month=?",
            (book_code, year_month)
        )
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'ok': False, 'error': '该月未锁定'}), 404
    finally:
        conn.close()
    log_action(None, 'unlock', session['username'],
               user.get('display_name', session['username']),
               f"账套={book_code} 月份={year_month} 备注={note}")
    _broadcast({'event': 'unlock', 'book_code': book_code, 'year_month': year_month})
    return jsonify({'ok': True})


# ============ AI 配置 API（仅 admin） ============
@app.route('/api/ai_config', methods=['GET'])
@admin_required
def api_get_ai_config():
    cfg = _load_ai_config()
    # 返回时把密钥脱敏
    masked = cfg.copy()
    if masked.get('apikey'):
        k = masked['apikey']
        masked['apikey_masked'] = k[:6] + '****' + k[-4:] if len(k) > 12 else '****'
    masked.pop('apikey', None)
    return jsonify({'ok': True, 'config': masked})


@app.route('/api/ai_config', methods=['POST'])
@admin_required
def api_set_ai_config():
    data = request.get_json(force=True, silent=True) or {}
    provider = data.get('provider') or 'deepseek'
    apikey = (data.get('apikey') or '').strip()
    if provider not in ('deepseek', 'qwen', 'rule'):
        return jsonify({'ok': False, 'error': 'provider 非法'}), 400
    with _ai_config_lock:
        cfg = _load_ai_config()
        cfg['provider'] = provider
        if apikey:  # 空字符串保留旧 key
            cfg['apikey'] = apikey
        _save_ai_config(cfg)
    return jsonify({'ok': True})


# ============ AI / 规则解析 ============
RULE_DATE_MAP = {
    '今天': 0, '今儿': 0, '今日': 0,
    '昨天': -1, '昨日': -1, '昨儿': -1,
    '前天': -2, '大前天': -3,
}

INCOME_KEYWORDS = ['收入', '工资', '报销', '到账', '收到', '转来', '发了', '奖金',
                   '分红', '退款', '卖了', '卖出', '回款']
EXPENSE_KEYWORDS = ['花', '买', '付', '支出', '花费', '给', '打车', '吃', '消费',
                    '交', '缴', '转账给', '充', '订']


def rule_parse(text):
    """规则解析作为兜底。"""
    # 日期
    parsed_date = None
    for key, offset in RULE_DATE_MAP.items():
        if key in text:
            d = datetime.date.today() + datetime.timedelta(days=offset)
            parsed_date = d
            break
    md = re.search(r'(\d{1,2})月(\d{1,2})[日号]', text)
    if md:
        now = datetime.date.today()
        parsed_date = datetime.date(now.year, int(md.group(1)), int(md.group(2)))
    if parsed_date is None:
        parsed_date = datetime.date.today()

    # 金额
    amount = 0.0
    m = re.search(r'(\d+\.?\d*)\s*[块元]?', text)
    if m:
        try:
            amount = float(m.group(1))
        except ValueError:
            amount = 0.0

    # 收支
    typ = 'expense'
    for k in INCOME_KEYWORDS:
        if k in text:
            typ = 'income'
            break
    if typ == 'expense':
        for k in EXPENSE_KEYWORDS:
            if k in text:
                typ = 'expense'
                break

    # 摘要
    summary = text
    for key in RULE_DATE_MAP:
        summary = summary.replace(key, '')
    summary = re.sub(r'\d{1,2}月\d{1,2}[日号]', '', summary)
    summary = re.sub(r'\d+\.?\d*\s*[块元钱]?', '', summary)
    summary = re.sub(r'[了的呢吧啊嗯，。！？、\s]+', ' ', summary).strip()
    if not summary:
        summary = '收入' if typ == 'income' else '支出'

    return {
        'date': f"{parsed_date.year}/{parsed_date.month}/{parsed_date.day}",
        'summary': summary[:30],
        'type': typ,
        'amount': amount,
    }


def ai_parse(text, provider, apikey, categories=None):
    """调用 DeepSeek / 通义千问 解析。categories 为可选类别列表，AI 会从中选一个。"""
    import urllib.request
    import urllib.error
    today = datetime.date.today()
    today_str = f"{today.year}/{today.month}/{today.day}"

    cat_part = ''
    if categories:
        cat_str = '、'.join(f'"{c}"' for c in categories)
        cat_part = f"""
- category: 类别，必须从下面列表中选一个：{cat_str}。判断不出来就选"其他"。"""

    prompt = f"""你是一个财务记账助手。请从用户的一句话中提取财务流水信息，严格输出JSON格式，不要任何额外文字。

今天的日期是 {today_str}。

需要提取的字段：
- date: 日期，格式 YYYY/M/D（注意月日不补零），如果用户没说日期就用今天
- summary: 摘要，简洁描述这笔流水（例如"午饭"、"打车"、"工资"），不要包含金额和日期
- type: "income"（收入）或 "expense"（支出）
- amount: 金额数字（人民币元）{cat_part}

用户说: "{text}"

只输出JSON：{{"date":"...","summary":"...","type":"...","amount":数字{', "category":"..."' if categories else ''}}}"""

    if provider == 'deepseek':
        url = 'https://api.deepseek.com/chat/completions'
        body = {
            'model': 'deepseek-chat',
            'messages': [{'role': 'user', 'content': prompt}],
            'temperature': 0,
            'response_format': {'type': 'json_object'},
        }
    elif provider == 'qwen':
        url = 'https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions'
        body = {
            'model': 'qwen-turbo',
            'messages': [{'role': 'user', 'content': prompt}],
            'temperature': 0,
        }
    else:
        raise ValueError(f'未知 provider: {provider}')

    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode('utf-8'),
        headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {apikey}',
        },
        method='POST',
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        result = json.loads(resp.read().decode('utf-8'))
    content = result['choices'][0]['message']['content']
    m = re.search(r'\{[\s\S]*\}', content)
    if not m:
        raise ValueError('AI 返回格式错误')
    parsed = json.loads(m.group(0))
    out = {
        'date': parsed.get('date') or today_str,
        'summary': str(parsed.get('summary') or '')[:30],
        'type': parsed.get('type') if parsed.get('type') in ('income', 'expense') else 'expense',
        'amount': float(parsed.get('amount') or 0),
    }
    if categories:
        cat = parsed.get('category') or ''
        if cat not in categories:
            cat = '其他' if '其他' in categories else (categories[0] if categories else '')
        out['category'] = cat
    return out


@app.route('/api/parse', methods=['POST'])
@login_required
def api_parse():
    data = request.get_json(force=True, silent=True) or {}
    text = (data.get('text') or '').strip()
    book_code = (data.get('book_code') or 'default').strip()
    if not text:
        return jsonify({'ok': False, 'error': '文本为空'}), 400
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403

    # 读取该账套的类别清单，提供给 AI
    categories_list = []
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT name FROM categories WHERE book_code=? AND enabled=1 "
            "ORDER BY sort_order ASC", (book_code,)
        ).fetchall()
        categories_list = [r['name'] for r in rows]
    finally:
        conn.close()

    cfg = _load_ai_config()
    provider = cfg.get('provider', 'deepseek')
    apikey = cfg.get('apikey', '')

    parsed = None
    used = 'rule'
    error_msg = None
    if provider in ('deepseek', 'qwen') and apikey:
        try:
            parsed = ai_parse(text, provider, apikey, categories_list)
            used = provider
        except Exception as e:
            error_msg = str(e)
            print(f'[ai_parse] 失败，降级到 rule: {e}')
    if parsed is None:
        parsed = rule_parse(text)
        used = 'rule'
        # 规则模式给一个默认类别
        if categories_list:
            parsed['category'] = '其他' if '其他' in categories_list else categories_list[0]

    return jsonify({
        'ok': True,
        'parsed': parsed,
        'used': used,
        'raw_text': text,
        'ai_error': error_msg,
    })


# ============ 语音识别 ASR（火山引擎 豆包大模型 录音文件识别） ============
ASR_CONFIG_FILE = os.path.join(BASE_DIR, 'asr_config.json')


def _load_asr_config():
    try:
        with open(ASR_CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def volc_asr_recognize(audio_bytes, audio_format, api_key, resource_id='volc.seedasr.auc'):
    """调用火山引擎大模型录音文件识别：submit 提交（音频 base64 内联）-> query 轮询取结果。"""
    import urllib.request
    import urllib.error
    import base64
    import uuid

    submit_url = 'https://openspeech.bytedance.com/api/v3/auc/bigmodel/submit'
    query_url = 'https://openspeech.bytedance.com/api/v3/auc/bigmodel/query'
    request_id = str(uuid.uuid4())
    audio_b64 = base64.b64encode(audio_bytes).decode('ascii')

    submit_body = {
        'user': {'uid': 'finance_app'},
        'audio': {'format': audio_format, 'data': audio_b64},
        'request': {
            'model_name': 'bigmodel',
            'enable_itn': True,
            'enable_punc': False,
            'enable_ddc': False,
            'show_utterances': True,
        },
    }
    submit_headers = {
        'Content-Type': 'application/json',
        'x-api-key': api_key,
        'X-Api-Resource-Id': resource_id,
        'X-Api-Request-Id': request_id,
        'X-Api-Sequence': '-1',
    }
    req = urllib.request.Request(
        submit_url, data=json.dumps(submit_body).encode('utf-8'),
        headers=submit_headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            submit_status = resp.headers.get('X-Api-Status-Code')
            resp.read()
    except urllib.error.HTTPError as e:
        raise ValueError(f'火山提交HTTP错误 {e.code}: {e.read().decode("utf-8", "ignore")[:200]}')
    if submit_status not in (None, '20000000'):
        raise ValueError(f'火山提交失败 status={submit_status}')

    query_headers = {
        'Content-Type': 'application/json',
        'x-api-key': api_key,
        'X-Api-Resource-Id': resource_id,
        'X-Api-Request-Id': request_id,
    }
    for _ in range(30):  # 最多约 30 秒
        qreq = urllib.request.Request(
            query_url, data=b'{}', headers=query_headers, method='POST')
        with urllib.request.urlopen(qreq, timeout=20) as qresp:
            qstatus = qresp.headers.get('X-Api-Status-Code')
            qbody = qresp.read().decode('utf-8')
        if qstatus == '20000000':  # 完成
            data = json.loads(qbody or '{}')
            result = data.get('result') or {}
            text = result.get('text')
            if not text and isinstance(result.get('utterances'), list):
                text = ''.join(u.get('text', '') for u in result['utterances'])
            return (text or '').strip()
        if qstatus in ('20000001', '20000002'):  # 处理中/排队中
            time.sleep(1)
            continue
        raise ValueError(f'火山识别失败 status={qstatus} body={qbody[:200]}')
    raise ValueError('火山识别超时')


@app.route('/api/asr', methods=['POST'])
@login_required
def api_asr():
    cfg = _load_asr_config()
    api_key = (cfg.get('api_key') or '').strip()
    resource_id = (cfg.get('resource_id') or 'volc.seedasr.auc').strip()
    if not api_key:
        return jsonify({'ok': False, 'error': '未配置火山语音识别 Key（asr_config.json）'}), 400
    audio_bytes = request.get_data()
    if not audio_bytes:
        return jsonify({'ok': False, 'error': '没有收到音频'}), 400
    audio_format = (request.args.get('format') or 'wav').strip()
    try:
        text = volc_asr_recognize(audio_bytes, audio_format, api_key, resource_id)
    except Exception as e:
        print(f'[asr] 火山识别失败: {e}')
        return jsonify({'ok': False, 'error': f'语音识别失败: {e}'}), 500
    if not text:
        return jsonify({'ok': False, 'error': '没有识别到内容，请说清楚一点'}), 200
    return jsonify({'ok': True, 'text': text})


# ============ 流水 CRUD ============
def _entry_to_dict(row):
    return {
        'id': row['id'],
        'date': row['date'],
        'year_month': row['year_month'],
        'summary': row['summary'],
        'type': row['type'],
        'amount': row['amount'],
        'remark': row['remark'] or '',
        'category': row['category'] if 'category' in row.keys() else '',
        'book_code': row['book_code'] if 'book_code' in row.keys() else 'default',
        'creator': row['creator'],
        'creator_name': row['creator_name'],
        'created_at': row['created_at'],
        'updated_at': row['updated_at'],
    }


def _normalize_date(s):
    """支持 2026/5/24 / 2026-5-24 / 2026/05/24，输出 (YYYY/M/D, YYYY-MM)。"""
    s = (s or '').strip().replace('-', '/')
    parts = s.split('/')
    if len(parts) != 3:
        raise ValueError(f'日期格式错误: {s}')
    y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
    datetime.date(y, m, d)  # 校验合法
    return f"{y}/{m}/{d}", f"{y:04d}-{m:02d}"


@app.route('/api/entries', methods=['GET'])
@login_required
def api_list_entries():
    """列表查询，支持多种过滤：
    - book_code              账套代号（必传或默认 default）
    - year_month=YYYY-MM     仅查某月
    - date_from=YYYY/M/D     起始日期
    - date_to=YYYY/M/D       结束日期
    - creator=username       按录入人
    - type=income|expense    按类型
    - category=name          按类别
    - q=keyword              文字关键字（摘要/备注/原始语音文字）
    - amount_min / amount_max 金额范围
    - limit                  最多返回条数（默认 2000）
    """
    book_code = (request.args.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403

    ym = request.args.get('year_month')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    creator = request.args.get('creator')
    typ = request.args.get('type')
    category = request.args.get('category')
    keyword = (request.args.get('q') or '').strip()
    amount_min = request.args.get('amount_min')
    amount_max = request.args.get('amount_max')
    try:
        limit = int(request.args.get('limit') or 2000)
        limit = max(1, min(limit, 10000))
    except ValueError:
        limit = 2000

    df_obj = None
    dt_obj = None
    try:
        if date_from:
            df_str, _ = _normalize_date(date_from)
            df_obj = datetime.date(*[int(x) for x in df_str.split('/')])
        if date_to:
            dt_str, _ = _normalize_date(date_to)
            dt_obj = datetime.date(*[int(x) for x in dt_str.split('/')])
    except ValueError as e:
        return jsonify({'ok': False, 'error': f'日期格式错: {e}'}), 400

    sql = "SELECT * FROM entries WHERE deleted=0 AND book_code=?"
    args = [book_code]
    if ym:
        sql += " AND year_month=?"
        args.append(ym)
    if creator:
        sql += " AND creator=?"
        args.append(creator)
    if typ in ('income', 'expense'):
        sql += " AND type=?"
        args.append(typ)
    if category:
        sql += " AND category=?"
        args.append(category)
    if keyword:
        sql += " AND (summary LIKE ? OR remark LIKE ? OR raw_text LIKE ?)"
        kw = f'%{keyword}%'
        args.extend([kw, kw, kw])
    if amount_min is not None and amount_min != '':
        try:
            sql += " AND amount >= ?"
            args.append(float(amount_min))
        except ValueError:
            return jsonify({'ok': False, 'error': 'amount_min 不是数字'}), 400
    if amount_max is not None and amount_max != '':
        try:
            sql += " AND amount <= ?"
            args.append(float(amount_max))
        except ValueError:
            return jsonify({'ok': False, 'error': 'amount_max 不是数字'}), 400
    sql += f" ORDER BY date DESC, created_at DESC LIMIT {limit}"

    conn = get_db()
    try:
        rows = conn.execute(sql, args).fetchall()
        entries = [_entry_to_dict(r) for r in rows]

        if df_obj or dt_obj:
            def in_range(e):
                d = datetime.date(*[int(x) for x in e['date'].split('/')])
                if df_obj and d < df_obj:
                    return False
                if dt_obj and d > dt_obj:
                    return False
                return True
            entries = [e for e in entries if in_range(e)]

        income = sum(e['amount'] for e in entries if e['type'] == 'income')
        expense = sum(e['amount'] for e in entries if e['type'] == 'expense')

        # 顺便返回当前月份的锁定状态（如果是按月查询）
        lock_info = None
        if ym:
            lk = get_month_lock(book_code, ym)
            if lk:
                lock_info = {
                    'locked': True,
                    'locked_by_name': lk['locked_name'],
                    'locked_at': lk['locked_at'],
                    'note': lk['note'],
                }
            else:
                lock_info = {'locked': False}

        return jsonify({
            'ok': True,
            'entries': entries,
            'summary': {
                'count': len(entries),
                'income': round(income, 2),
                'expense': round(expense, 2),
            },
            'lock_info': lock_info,
            'book_code': book_code,
        })
    finally:
        conn.close()


@app.route('/api/entries', methods=['POST'])
@login_required
def api_create_entry():
    data = request.get_json(force=True, silent=True) or {}
    book_code = (data.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403
    try:
        date, ym = _normalize_date(data.get('date'))
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    # 锁定检查
    if is_month_locked(book_code, ym):
        return jsonify({
            'ok': False,
            'error': f'{ym} 已锁定，不能新增记录。如需修改请管理员先解锁。',
            'locked': True,
        }), 423  # 423 Locked

    summary = (data.get('summary') or '').strip()
    typ = data.get('type')
    category = (data.get('category') or '').strip()
    try:
        amount = float(data.get('amount') or 0)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': '金额格式错误'}), 400
    remark = (data.get('remark') or '').strip()
    raw_text = (data.get('raw_text') or '').strip()

    if not summary:
        return jsonify({'ok': False, 'error': '摘要不能为空'}), 400
    if typ not in ('income', 'expense'):
        return jsonify({'ok': False, 'error': 'type 非法'}), 400
    if amount <= 0:
        return jsonify({'ok': False, 'error': '金额必须大于 0'}), 400

    user = get_user(session['username'])
    creator = session['username']
    creator_name = user.get('display_name', creator)
    ts = now_str()

    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO entries (book_code, date, year_month, summary, type, amount, "
            "category, remark, raw_text, creator, creator_name, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (book_code, date, ym, summary, typ, amount, category, remark, raw_text,
             creator, creator_name, ts, ts)
        )
        conn.commit()
        entry_id = cur.lastrowid
        row = conn.execute("SELECT * FROM entries WHERE id=?", (entry_id,)).fetchone()
    finally:
        conn.close()

    log_action(entry_id, 'create', creator, creator_name,
               f"[{book_code}] {date} {summary} {'+' if typ=='income' else '-'}{amount}")

    entry = _entry_to_dict(row)
    _broadcast({'event': 'create', 'entry': entry, 'book_code': book_code})
    return jsonify({'ok': True, 'entry': entry})


@app.route('/api/entries/<int:entry_id>', methods=['PATCH'])
@login_required
def api_update_entry(entry_id):
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM entries WHERE id=? AND deleted=0",
                           (entry_id,)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': '记录不存在'}), 404

        book_code = row['book_code']
        if not can_access_book(session['username'], book_code):
            return jsonify({'ok': False, 'error': '无权访问该账套'}), 403

        # 权限：本人 或 admin 可改
        user = get_user(session['username'])
        if row['creator'] != session['username'] and user.get('role') != 'admin':
            return jsonify({'ok': False, 'error': '只能修改自己录入的记录'}), 403

        # 锁定检查：原月份或目标月份任一被锁，都拒绝
        # admin 也不能直接改锁定月份，必须先解锁
        if is_month_locked(book_code, row['year_month']):
            return jsonify({
                'ok': False,
                'error': f'{row["year_month"]} 已锁定，请先解锁该月份。',
                'locked': True,
            }), 423

        # 构造更新字段
        new_date = row['date']
        new_ym = row['year_month']
        if 'date' in data:
            try:
                new_date, new_ym = _normalize_date(data['date'])
            except ValueError as e:
                return jsonify({'ok': False, 'error': str(e)}), 400
            # 如果改到了不同月份，也要检查新月份是否锁定
            if new_ym != row['year_month'] and is_month_locked(book_code, new_ym):
                return jsonify({
                    'ok': False,
                    'error': f'目标月份 {new_ym} 已锁定，无法移入。',
                    'locked': True,
                }), 423

        new_summary = (data.get('summary') or row['summary']).strip()
        new_type = data.get('type', row['type'])
        if new_type not in ('income', 'expense'):
            return jsonify({'ok': False, 'error': 'type 非法'}), 400
        try:
            new_amount = float(data['amount']) if 'amount' in data else row['amount']
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': '金额格式错误'}), 400
        if new_amount <= 0:
            return jsonify({'ok': False, 'error': '金额必须大于 0'}), 400
        new_remark = data.get('remark', row['remark']) or ''
        new_category = data.get('category', row['category']) if 'category' in data else (row['category'] or '')
        ts = now_str()

        conn.execute(
            "UPDATE entries SET date=?, year_month=?, summary=?, type=?, amount=?, "
            "category=?, remark=?, updated_at=? WHERE id=?",
            (new_date, new_ym, new_summary, new_type, new_amount,
             new_category, new_remark, ts, entry_id)
        )
        conn.commit()
        row2 = conn.execute("SELECT * FROM entries WHERE id=?", (entry_id,)).fetchone()
    finally:
        conn.close()

    log_action(entry_id, 'update', session['username'],
               get_user(session['username']).get('display_name', session['username']),
               f"[{book_code}] {new_date} {new_summary} {new_amount}")
    entry = _entry_to_dict(row2)
    _broadcast({'event': 'update', 'entry': entry, 'book_code': book_code})
    return jsonify({'ok': True, 'entry': entry})


@app.route('/api/entries/<int:entry_id>', methods=['DELETE'])
@login_required
def api_delete_entry(entry_id):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM entries WHERE id=? AND deleted=0",
                           (entry_id,)).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': '记录不存在'}), 404

        book_code = row['book_code']
        if not can_access_book(session['username'], book_code):
            return jsonify({'ok': False, 'error': '无权访问该账套'}), 403

        user = get_user(session['username'])
        if row['creator'] != session['username'] and user.get('role') != 'admin':
            return jsonify({'ok': False, 'error': '只能删除自己录入的记录'}), 403

        # 锁定检查
        if is_month_locked(book_code, row['year_month']):
            return jsonify({
                'ok': False,
                'error': f'{row["year_month"]} 已锁定，请先解锁该月份。',
                'locked': True,
            }), 423

        # 软删除
        conn.execute("UPDATE entries SET deleted=1, updated_at=? WHERE id=?",
                     (now_str(), entry_id))
        conn.commit()
    finally:
        conn.close()
    log_action(entry_id, 'delete', session['username'],
               get_user(session['username']).get('display_name', session['username']),
               f"[{book_code}] {row['date']} {row['summary']}")
    _broadcast({'event': 'delete', 'entry_id': entry_id, 'book_code': book_code})
    return jsonify({'ok': True})


# ============ 导出 Excel ============
@app.route('/api/export', methods=['GET'])
@login_required
def api_export():
    """导出 Excel，列结构对齐用户提供的模板。
    参数：
      - book_code              账套代号（必传或默认 default）
      - year_month=YYYY-MM     按月导出（最常用）
      - date_from/date_to/q/creator/type/category/amount_min/amount_max  按搜索条件导出
      - include_creator=1      额外含「录入人」列
      - include_category=1     额外含「类别」列
    """
    book_code = (request.args.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403
    book = get_book(book_code)

    ym = request.args.get('year_month')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    creator = request.args.get('creator')
    typ = request.args.get('type')
    category = request.args.get('category')
    keyword = (request.args.get('q') or '').strip()
    amount_min = request.args.get('amount_min')
    amount_max = request.args.get('amount_max')
    include_creator = request.args.get('include_creator') == '1'
    include_category = request.args.get('include_category') == '1'

    if ym and not re.match(r'^\d{4}-\d{2}$', ym):
        return jsonify({'ok': False, 'error': 'year_month 格式错'}), 400

    # 日期范围预处理
    df_obj = None
    dt_obj = None
    try:
        if date_from:
            df_str, _ = _normalize_date(date_from)
            df_obj = datetime.date(*[int(x) for x in df_str.split('/')])
        if date_to:
            dt_str, _ = _normalize_date(date_to)
            dt_obj = datetime.date(*[int(x) for x in dt_str.split('/')])
    except ValueError as e:
        return jsonify({'ok': False, 'error': f'日期格式错: {e}'}), 400

    sql = "SELECT * FROM entries WHERE deleted=0 AND book_code=?"
    args = [book_code]
    if ym:
        sql += " AND year_month=?"
        args.append(ym)
    if creator:
        sql += " AND creator=?"
        args.append(creator)
    if typ in ('income', 'expense'):
        sql += " AND type=?"
        args.append(typ)
    if category:
        sql += " AND category=?"
        args.append(category)
    if keyword:
        sql += " AND (summary LIKE ? OR remark LIKE ? OR raw_text LIKE ?)"
        kw = f'%{keyword}%'
        args.extend([kw, kw, kw])
    if amount_min is not None and amount_min != '':
        try:
            sql += " AND amount >= ?"
            args.append(float(amount_min))
        except ValueError:
            return jsonify({'ok': False, 'error': 'amount_min 不是数字'}), 400
    if amount_max is not None and amount_max != '':
        try:
            sql += " AND amount <= ?"
            args.append(float(amount_max))
        except ValueError:
            return jsonify({'ok': False, 'error': 'amount_max 不是数字'}), 400
    sql += " ORDER BY date ASC, created_at ASC"

    conn = get_db()
    try:
        rows = conn.execute(sql, args).fetchall()
    finally:
        conn.close()

    # Python 端日期范围过滤
    if df_obj or dt_obj:
        filtered = []
        for r in rows:
            d = datetime.date(*[int(x) for x in r['date'].split('/')])
            if df_obj and d < df_obj:
                continue
            if dt_obj and d > dt_obj:
                continue
            filtered.append(r)
        rows = filtered

    if not rows:
        return jsonify({'ok': False, 'error': '没有匹配的数据可导出'}), 404

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return jsonify({'ok': False, 'error': '服务器缺少 openpyxl，请 pip install openpyxl'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 列头严格对齐用户模板
    headers = ['日期', '流水号', '摘要', '账户对方科目', '辅助核算', '对方户名',
               '对方账号', '对方银行', '收入', '支出', '备注']
    if include_category:
        headers.append('类别')
    if include_creator:
        headers.append('录入人')

    ws.append(headers)
    # 表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill('solid', start_color='F2F2F2')
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, row in enumerate(rows, 1):
        remark_parts = []
        if row['remark']:
            remark_parts.append(row['remark'])
        if row['raw_text'] and row['raw_text'] != row['summary']:
            remark_parts.append(f"[语音]{row['raw_text']}")
        remark = ' / '.join(remark_parts)

        data_row = [
            row['date'],
            i,
            row['summary'],
            '', '', '', '', '',
            row['amount'] if row['type'] == 'income' else '',
            row['amount'] if row['type'] == 'expense' else '',
            remark,
        ]
        if include_category:
            data_row.append(row['category'] or '')
        if include_creator:
            data_row.append(row['creator_name'])
        ws.append(data_row)
        for col_idx in range(1, len(data_row) + 1):
            ws.cell(row=i + 1, column=col_idx).border = border

    # 列宽
    widths = [12, 8, 24, 14, 12, 14, 18, 14, 12, 12, 24]
    if include_category:
        widths.append(12)
    if include_creator:
        widths.append(12)
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 文件名根据导出类型动态生成；带账套名
    book_part = book['name'] if book else book_code
    # 把账套名里特殊字符去掉
    safe_book = re.sub(r'[^\w\u4e00-\u9fff]', '_', book_part)[:20]
    if ym:
        y, m = ym.split('-')
        fname = f"财务日记账_{safe_book}_{y}年{m}月.xlsx"
        ascii_name = f"finance_{book_code}_{y}_{m}.xlsx"
    elif df_obj and dt_obj:
        fname = f"财务日记账_{safe_book}_{df_obj.strftime('%Y%m%d')}-{dt_obj.strftime('%Y%m%d')}.xlsx"
        ascii_name = f"finance_{book_code}_{df_obj.strftime('%Y%m%d')}-{dt_obj.strftime('%Y%m%d')}.xlsx"
    elif keyword:
        from datetime import datetime as _dt
        ts = _dt.now().strftime('%Y%m%d_%H%M%S')
        safe_kw = re.sub(r'[^\w\u4e00-\u9fff]', '_', keyword)[:30]
        fname = f"财务搜索_{safe_book}_{safe_kw}_{ts}.xlsx"
        ascii_name = f"finance_search_{book_code}_{ts}.xlsx"
    else:
        from datetime import datetime as _dt
        ts = _dt.now().strftime('%Y%m%d_%H%M%S')
        fname = f"财务日记账_{safe_book}_筛选导出_{ts}.xlsx"
        ascii_name = f"finance_filtered_{book_code}_{ts}.xlsx"

    from urllib.parse import quote
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = (
        f"attachment; filename={ascii_name}; "
        f"filename*=UTF-8''{quote(fname)}"
    )
    return resp


# ============ 导入 Excel ============
def _excel_date_to_str(val):
    """把 Excel 单元格里的日期值转成 YYYY/M/D 字符串。
    支持：
      - datetime.datetime / datetime.date 对象（openpyxl 自动转换的日期单元格）
      - 字符串 "2024/3/15" / "2024-3-15" / "2024/03/15" / "2024.3.15" / "2024年3月15日"
      - 字符串里有时分（"2024/3/15 10:00:00"）也能切掉时分
      - Excel 序列号（罕见，但有人手工填）—— 当作 1900-01-01 起的天数
    返回 (date_str, year_month) 或 抛 ValueError。
    """
    if val is None or val == '':
        raise ValueError('日期为空')
    # datetime 对象
    if isinstance(val, (datetime.datetime, datetime.date)):
        d = val.date() if isinstance(val, datetime.datetime) else val
        return f"{d.year}/{d.month}/{d.day}", f"{d.year:04d}-{d.month:02d}"
    # 数字（Excel 序列号）
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            # Excel 序列号：1900-01-01 = 1（注意 Excel bug：1900 被错当成闰年，差 1 天）
            base = datetime.date(1899, 12, 30)
            d = base + datetime.timedelta(days=int(val))
            return f"{d.year}/{d.month}/{d.day}", f"{d.year:04d}-{d.month:02d}"
        except Exception:
            raise ValueError(f'数字日期无法解析: {val}')
    # 字符串
    s = str(val).strip()
    if not s:
        raise ValueError('日期为空')
    # 切掉时分
    s = re.split(r'[\s　]', s)[0]
    # 统一分隔符
    s = s.replace('年', '/').replace('月', '/').replace('日', '').replace('.', '/').replace('-', '/')
    s = s.strip('/')
    parts = s.split('/')
    if len(parts) != 3:
        raise ValueError(f'日期格式无法识别: {val}')
    try:
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        # 防呆：2 位年份猜成 20xx
        if y < 100:
            y += 2000
        datetime.date(y, m, d)  # 校验合法
        return f"{y}/{m}/{d}", f"{y:04d}-{m:02d}"
    except (ValueError, TypeError):
        raise ValueError(f'日期格式无法识别: {val}')


def _parse_excel_amount(val):
    """把 Excel 单元格的金额转成 float。支持空、字符串、带逗号/¥/￥。"""
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    # 去掉常见符号
    s = s.replace(',', '').replace('¥', '').replace('￥', '').replace(' ', '').replace('元', '').replace('块', '')
    try:
        return float(s)
    except ValueError:
        return 0.0


@app.route('/api/import', methods=['POST'])
@login_required
def api_import():
    """导入历史 Excel。
    multipart/form-data:
        file: .xlsx 文件
        target_user: 把所有记录归到这个用户名下（必须是已存在的用户）
                     不传或空字符串则归在当前登录用户下
        book_code:   导入到哪个账套（默认 default）
    返回：{ok, imported, skipped, errors: [{row, reason}], preview: [...]}
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '未上传文件'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': '文件名为空'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'ok': False, 'error': '只支持 .xlsx / .xlsm 文件'}), 400

    book_code = (request.form.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403
    if not get_book(book_code):
        return jsonify({'ok': False, 'error': '账套不存在'}), 404

    target_user = (request.form.get('target_user') or '').strip()
    if target_user:
        target = get_user(target_user)
        if not target:
            return jsonify({'ok': False, 'error': f'目标用户不存在: {target_user}'}), 400
        target_username = target_user
        target_display = target.get('display_name', target_user)
    else:
        target_username = session['username']
        me = get_user(target_username)
        target_display = me.get('display_name', target_username)

    try:
        import openpyxl
    except ImportError:
        return jsonify({'ok': False, 'error': '服务器缺少 openpyxl'}), 500

    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'打开 Excel 失败: {e}'}), 400

    ws = wb.active

    # 找表头行：扫描前 5 行，找到包含"日期"/"摘要"/"收入"/"支出"等关键字的行
    HEADER_ALIASES = {
        'date':    ['日期', '日 期', '记账日期', '业务日期'],
        'no':      ['流水号', '编号', '序号', 'No', 'NO', '#'],
        'summary': ['摘要', '内容', '事项', '说明', '业务摘要'],
        'income':  ['收入', '贷方金额', '收方', '收入金额', '存入'],
        'expense': ['支出', '借方金额', '付方', '支出金额', '支取'],
        'remark':  ['备注', '说明', '附注'],
    }

    def _norm_header(s):
        if s is None:
            return ''
        return str(s).strip().replace(' ', '')

    col_map = {}  # 字段名 -> 列索引（0-based）
    header_row_idx = -1
    rows_iter = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    for ridx, row in enumerate(rows_iter):
        normalized = [_norm_header(c) for c in row]
        local_map = {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                a = alias.replace(' ', '')
                if a in normalized:
                    local_map[field] = normalized.index(a)
                    break
        # 至少要有 日期、摘要、收入、支出 才算合法表头
        if 'date' in local_map and 'summary' in local_map and \
           ('income' in local_map or 'expense' in local_map):
            col_map = local_map
            header_row_idx = ridx
            break

    if header_row_idx < 0:
        return jsonify({
            'ok': False,
            'error': '无法识别表头：前 5 行没找到「日期/摘要/收入/支出」列。请确认文件格式与模板一致'
        }), 400

    # 逐行解析
    imported = 0
    skipped = 0
    errors = []
    preview = []
    ts = now_str()

    conn = get_db()
    try:
        conn.execute('BEGIN')
        # 数据从 header_row_idx+1 行开始（0-based）→ Excel 行号是 header_row_idx+2
        for excel_row_num, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 2, values_only=True),
            start=header_row_idx + 2
        ):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue  # 跳过空行

            try:
                date_val = row[col_map['date']] if col_map['date'] < len(row) else None
                summary_val = row[col_map['summary']] if col_map['summary'] < len(row) else None
                income_val = row[col_map['income']] if 'income' in col_map and col_map['income'] < len(row) else None
                expense_val = row[col_map['expense']] if 'expense' in col_map and col_map['expense'] < len(row) else None
                remark_val = row[col_map['remark']] if 'remark' in col_map and col_map['remark'] < len(row) else None

                # 日期
                try:
                    date_str, year_month = _excel_date_to_str(date_val)
                except ValueError as e:
                    errors.append({'row': excel_row_num, 'reason': str(e)})
                    skipped += 1
                    continue

                # 摘要
                summary = str(summary_val or '').strip()
                if not summary:
                    errors.append({'row': excel_row_num, 'reason': '摘要为空'})
                    skipped += 1
                    continue

                # 金额：收入 vs 支出
                income_amt = _parse_excel_amount(income_val)
                expense_amt = _parse_excel_amount(expense_val)
                if income_amt > 0 and expense_amt > 0:
                    # 两列都有钱，按金额大的优先（罕见情况）
                    if income_amt >= expense_amt:
                        typ = 'income'
                        amount = income_amt
                    else:
                        typ = 'expense'
                        amount = expense_amt
                elif income_amt > 0:
                    typ = 'income'
                    amount = income_amt
                elif expense_amt > 0:
                    typ = 'expense'
                    amount = expense_amt
                else:
                    errors.append({'row': excel_row_num, 'reason': '收入和支出都为 0 或为空'})
                    skipped += 1
                    continue

                remark = str(remark_val or '').strip()

                # 锁定月份保护：跳过
                if is_month_locked(book_code, year_month):
                    errors.append({
                        'row': excel_row_num,
                        'reason': f'{year_month} 已锁定，跳过'
                    })
                    skipped += 1
                    continue

                # 插入（用户选了"重复也导入"，所以不做去重）
                conn.execute(
                    "INSERT INTO entries (book_code, date, year_month, summary, type, "
                    "amount, remark, raw_text, creator, creator_name, "
                    "created_at, updated_at) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (book_code, date_str, year_month, summary, typ, amount,
                     remark, '', target_username, target_display, ts, ts)
                )
                imported += 1
                if len(preview) < 5:
                    preview.append({
                        'row': excel_row_num,
                        'date': date_str,
                        'summary': summary,
                        'type': typ,
                        'amount': amount,
                    })
            except Exception as e:
                errors.append({'row': excel_row_num, 'reason': f'处理异常: {e}'})
                skipped += 1
                continue

        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': f'导入事务失败: {e}'}), 500
    finally:
        conn.close()

    log_action(None, 'import', session['username'],
               get_user(session['username']).get('display_name', session['username']),
               f"[{book_code}] 导入 {imported} 条到 {target_display}，跳过 {skipped} 条")

    # 广播一次"刷新"事件
    _broadcast({'event': 'import', 'imported': imported,
                'target': target_display, 'book_code': book_code})

    return jsonify({
        'ok': True,
        'imported': imported,
        'skipped': skipped,
        'book_code': book_code,
        'target_user': target_username,
        'target_display': target_display,
        'errors': errors[:50],  # 最多返回 50 条错误
        'preview': preview,
    })


# ============ 月份概览 ============
@app.route('/api/months_overview', methods=['GET'])
@login_required
def api_months_overview():
    """返回某账套下有数据的所有月份，按降序，附带锁定状态。"""
    book_code = (request.args.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT year_month, COUNT(*) as cnt, "
            "  SUM(CASE WHEN type='income' THEN amount ELSE 0 END) as income, "
            "  SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense "
            "FROM entries WHERE deleted=0 AND book_code=? "
            "GROUP BY year_month ORDER BY year_month DESC",
            (book_code,)
        ).fetchall()
        # 查锁定情况
        lock_rows = conn.execute(
            "SELECT year_month FROM month_locks WHERE book_code=?",
            (book_code,)
        ).fetchall()
        locked_set = set(r['year_month'] for r in lock_rows)
        months = [{
            'year_month': r['year_month'],
            'count': r['cnt'],
            'income': round(r['income'] or 0, 2),
            'expense': round(r['expense'] or 0, 2),
            'locked': r['year_month'] in locked_set,
        } for r in rows]
        return jsonify({'ok': True, 'months': months, 'book_code': book_code})
    finally:
        conn.close()


# ============ 统计 ============
@app.route('/api/stats/monthly_trend', methods=['GET'])
@login_required
def api_stats_monthly_trend():
    """最近 N 个月（默认 12）的月度收支趋势。"""
    book_code = (request.args.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403
    try:
        n = int(request.args.get('months') or 12)
        n = max(1, min(n, 60))
    except ValueError:
        n = 12

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT year_month, "
            "  SUM(CASE WHEN type='income' THEN amount ELSE 0 END) as income, "
            "  SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense "
            "FROM entries WHERE deleted=0 AND book_code=? "
            "GROUP BY year_month ORDER BY year_month DESC LIMIT ?",
            (book_code, n)
        ).fetchall()
        items = [{
            'year_month': r['year_month'],
            'income': round(r['income'] or 0, 2),
            'expense': round(r['expense'] or 0, 2),
        } for r in rows]
        items.reverse()  # 时间正序便于画图
        return jsonify({'ok': True, 'items': items})
    finally:
        conn.close()


@app.route('/api/stats/category', methods=['GET'])
@login_required
def api_stats_category():
    """按类别统计（支持日期范围/月份过滤）。返回收入和支出两个分组。"""
    book_code = (request.args.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403

    ym = request.args.get('year_month')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    df_obj = None
    dt_obj = None
    try:
        if date_from:
            df_str, _ = _normalize_date(date_from)
            df_obj = datetime.date(*[int(x) for x in df_str.split('/')])
        if date_to:
            dt_str, _ = _normalize_date(date_to)
            dt_obj = datetime.date(*[int(x) for x in dt_str.split('/')])
    except ValueError as e:
        return jsonify({'ok': False, 'error': f'日期格式错: {e}'}), 400

    sql = "SELECT date, type, amount, category FROM entries WHERE deleted=0 AND book_code=?"
    args = [book_code]
    if ym:
        sql += " AND year_month=?"
        args.append(ym)

    conn = get_db()
    try:
        rows = conn.execute(sql, args).fetchall()
    finally:
        conn.close()

    income_cats = {}
    expense_cats = {}
    income_total = 0.0
    expense_total = 0.0
    for r in rows:
        d = datetime.date(*[int(x) for x in r['date'].split('/')])
        if df_obj and d < df_obj:
            continue
        if dt_obj and d > dt_obj:
            continue
        cat = r['category'] or '未分类'
        amt = r['amount']
        if r['type'] == 'income':
            income_cats[cat] = income_cats.get(cat, 0) + amt
            income_total += amt
        else:
            expense_cats[cat] = expense_cats.get(cat, 0) + amt
            expense_total += amt

    def to_list(d, total):
        items = sorted(d.items(), key=lambda x: -x[1])
        return [{
            'category': k,
            'amount': round(v, 2),
            'percent': round(v / total * 100, 1) if total > 0 else 0
        } for k, v in items]

    return jsonify({
        'ok': True,
        'income': {'items': to_list(income_cats, income_total), 'total': round(income_total, 2)},
        'expense': {'items': to_list(expense_cats, expense_total), 'total': round(expense_total, 2)},
    })


@app.route('/api/stats/range_summary', methods=['GET'])
@login_required
def api_stats_range_summary():
    """跨月汇总：自定义日期范围，按月分组的收支汇总 + 总计。"""
    book_code = (request.args.get('book_code') or 'default').strip()
    if not can_access_book(session['username'], book_code):
        return jsonify({'ok': False, 'error': '无权访问该账套'}), 403

    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    df_obj = None
    dt_obj = None
    try:
        if date_from:
            df_str, _ = _normalize_date(date_from)
            df_obj = datetime.date(*[int(x) for x in df_str.split('/')])
        if date_to:
            dt_str, _ = _normalize_date(date_to)
            dt_obj = datetime.date(*[int(x) for x in dt_str.split('/')])
    except ValueError as e:
        return jsonify({'ok': False, 'error': f'日期格式错: {e}'}), 400

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT year_month, date, type, amount FROM entries "
            "WHERE deleted=0 AND book_code=?",
            (book_code,)
        ).fetchall()
    finally:
        conn.close()

    months = {}
    income_total = 0.0
    expense_total = 0.0
    count_total = 0
    for r in rows:
        d = datetime.date(*[int(x) for x in r['date'].split('/')])
        if df_obj and d < df_obj:
            continue
        if dt_obj and d > dt_obj:
            continue
        ym = r['year_month']
        if ym not in months:
            months[ym] = {'income': 0.0, 'expense': 0.0, 'count': 0}
        amt = r['amount']
        months[ym]['count'] += 1
        count_total += 1
        if r['type'] == 'income':
            months[ym]['income'] += amt
            income_total += amt
        else:
            months[ym]['expense'] += amt
            expense_total += amt

    items = []
    for ym in sorted(months.keys()):
        m = months[ym]
        items.append({
            'year_month': ym,
            'income': round(m['income'], 2),
            'expense': round(m['expense'], 2),
            'net': round(m['income'] - m['expense'], 2),
            'count': m['count'],
        })

    return jsonify({
        'ok': True,
        'items': items,
        'totals': {
            'income': round(income_total, 2),
            'expense': round(expense_total, 2),
            'net': round(income_total - expense_total, 2),
            'count': count_total,
        },
    })


# ============ 实时同步（SSE） ============
_subscribers = []  # list of queue.Queue
_subscribers_lock = threading.Lock()


def _broadcast(event_data):
    """广播给所有订阅者。"""
    payload = json.dumps(event_data, ensure_ascii=False)
    with _subscribers_lock:
        dead = []
        for q in _subscribers:
            try:
                q.put_nowait(payload)
            except Exception:
                dead.append(q)
        for q in dead:
            try:
                _subscribers.remove(q)
            except ValueError:
                pass


@app.route('/api/stream')
@login_required
def api_stream():
    """Server-Sent Events，前端 EventSource 订阅。"""
    q = queue.Queue(maxsize=100)
    with _subscribers_lock:
        _subscribers.append(q)

    def gen():
        # 启动握手
        yield "event: ready\ndata: {}\n\n"
        last_ping = time.time()
        try:
            while True:
                try:
                    payload = q.get(timeout=15)
                    yield f"event: change\ndata: {payload}\n\n"
                except queue.Empty:
                    # 心跳，防代理切断
                    yield f": ping {int(time.time())}\n\n"
                    last_ping = time.time()
        except GeneratorExit:
            pass
        finally:
            with _subscribers_lock:
                try:
                    _subscribers.remove(q)
                except ValueError:
                    pass

    resp = Response(gen(), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'  # 关闭 Nginx 缓冲
    return resp


# ============ 启动 ============
if __name__ == '__main__':
    init_db()
    # 触发一次 _load_users 以确保默认 admin 已生成
    _load_users()
    _load_ai_config()
    print('=' * 60)
    print('  财务语音记账 - 多人协同服务启动中...')
    print('  端口: 5600')
    print('  默认管理员: admin / admin123（首次登录后请立即修改）')
    print('  数据库: finance.db')
    print('=' * 60)
    try:
        from waitress import serve
        print('  WSGI: waitress (threads=8)')
        print('=' * 60)
        # threads 不要太多，SSE 是长连接会占线程
        serve(app, host='0.0.0.0', port=5600,
              threads=16, ident='finance', expose_tracebacks=False,
              channel_timeout=600)
    except ImportError:
        print('  [警告] 未装 waitress，回退 Flask dev server')
        print('  建议: pip install waitress')
        print('=' * 60)
        app.run(host='0.0.0.0', port=5600, debug=False, threaded=True)
